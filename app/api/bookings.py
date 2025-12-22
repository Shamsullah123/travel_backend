from flask import Blueprint, request, jsonify, g, send_file
from app.middleware import token_required
from models.booking import Booking
from models.package import Package
from models.customer import Customer
from models.agency import Agency
from models.quotation import Quotation
from app.utils.serializers import mongo_to_dict
from app.utils.error_handlers import error_response, validation_error, not_found_error
from datetime import datetime
import json
import io
import csv
import traceback
import sys
from flask import make_response

bookings_bp = Blueprint('bookings', __name__)

@bookings_bp.route('', methods=['POST'])
@token_required
def create_booking():
    data = request.get_json()
    if not data:
        return error_response("No data provided", "INVALID_REQUEST")
    
    # Required fields for Direct Booking
    required_fields = ['customerId', 'packageId', 'category', 'totalAmount', 'finalAmount']
    for field in required_fields:
        if field not in data:
            return validation_error({field: 'Required field'})

    try:
        # Fetch References
        customer = Customer.objects(id=data['customerId'], agencyId=g.agency_id).first()
        if not customer:
            return not_found_error("Customer not found")

        package = Package.objects(id=data['packageId'], agencyId=g.agency_id).first()
        if not package:
            return not_found_error("Package not found")

        # Generate Booking Number (Simple Counter for MVP)
        # In prod, atomic counter is better
        count = Booking.objects(agencyId=g.agency_id).count() + 1
        booking_number = f"BK-{datetime.now().year}-{count:04d}"

        booking = Booking(
            agencyId=g.agency_id,
            customerId=customer,
            packageId=package,
            bookingNumber=booking_number,
            category=data['category'],
            baseAmount=data['totalAmount'], # Frontend sends base price as totalAmount
            discount=data.get('discount', 0),
            totalAmount=data['finalAmount'], # Frontend sends calculated final as finalAmount
            paidAmount=data.get('paidAmount', 0),
            status='Confirmed'
        )
        
        booking.save()
        
        return jsonify(mongo_to_dict(booking)), 201

    except Exception as e:
        return error_response(str(e), "SERVER_ERROR", 500)

@bookings_bp.route('', methods=['GET'])
@token_required
def get_bookings():
    # Helper to get bookings, possibly filtered by customerId
    customer_id = request.args.get('customerId')
    
    query = Booking.objects(agencyId=g.agency_id)
    if customer_id:
        query = query.filter(customerId=customer_id)
        
    bookings = query.order_by('-createdAt')
    data = mongo_to_dict(bookings)
    
    # Populate names for list view (MVP)
    for b_data in data:
        # Get booking object to access dereferenced fields or manually query
        # Using manual query here to avoid rereading all fields or relying on mongoengine dereference in loop
        if 'packageId' in b_data and b_data['packageId']:
            pkg = Package.objects(id=b_data['packageId'], agencyId=g.agency_id).first()
            if pkg: b_data['packageName'] = pkg.name
            
        if 'customerId' in b_data and b_data['customerId']:
            cust = Customer.objects(id=b_data['customerId'], agencyId=g.agency_id).first()
            if cust: b_data['customerName'] = cust.fullName

    return jsonify(data), 200

@bookings_bp.route('/export', methods=['GET'])
@token_required
def export_bookings():
    format_type = request.args.get('format', 'excel')
    
    # Fetch all bookings for the agency, including referenced documents
    bookings = Booking.objects(agencyId=g.agency_id).order_by('-createdAt')
    
    # Prepare comprehensive data
    booking_data = []
    print("DEBUG: Starting export generation")
    
    for booking in bookings:
        try:
            # Customer Details
            cust = booking.customerId
            # Safe access to customer fields in case of missing reference/deletion
            if cust:
                customer_info = {
                    'Customer Name': cust.fullName,
                    'Phone': cust.phone,
                    'CNIC': getattr(cust, 'cnic', ''),
                    'Passport Number': getattr(cust, 'passportNumber', ''),
                    'Finger Print': getattr(cust, 'finger_print', 'No'),
                    'Enrollment ID': getattr(cust, 'enrollment_id', ''),
                    'Gender': getattr(cust, 'gender', ''),
                    'Address': getattr(cust, 'address', '')
                }
            else:
                customer_info = {
                    'Customer Name': 'Unknown', 'Phone': '', 'CNIC': '', 
                    'Passport Number': '', 'Finger Print': 'No', 'Enrollment ID': '',
                    'Gender': '', 'Address': ''
                }
            
            # Package Details
            pkg = booking.packageId
            if pkg:
                package_info = {
                    'Package Name': pkg.name,
                    'Package Duration': pkg.duration,
                    'Start Date': pkg.startDate.strftime('%Y-%m-%d') if pkg.startDate else '',
                    'End Date': pkg.endDate.strftime('%Y-%m-%d') if pkg.endDate else '',
                    'Description': pkg.description
                }
            else:
                package_info = {
                    'Package Name': 'Direct Booking', 'Package Duration': '', 
                    'Start Date': '', 'End Date': '', 'Description': ''
                }
            
            # Booking Financials & Status
            booking_info = {
                'Booking Number': booking.bookingNumber,
                'Category': booking.category,
                'Status': booking.status,
                'Booking Date': booking.createdAt.strftime('%Y-%m-%d') if booking.createdAt else '',
                'Base Amount': float(booking.baseAmount) if booking.baseAmount else 0,
                'Discount': float(booking.discount) if booking.discount else 0,
                'Total Amount': float(booking.totalAmount) if booking.totalAmount else 0,
                'Paid Amount': float(booking.paidAmount) if booking.paidAmount else 0,
                'Balance Due': float(booking.balanceDue) if booking.balanceDue else 0,
                'PNR': getattr(booking, 'pnr', '') or '',
                'Supplier Ref': getattr(booking, 'supplierRef', '') or ''
            }
            
            # Combine all data
            row = {**booking_info, **customer_info, **package_info}
            booking_data.append(row)
            
        except Exception as inner_e:
            print(f"DEBUG: Error processing booking {booking.bookingNumber}: {str(inner_e)}")
            traceback.print_exc()
            # Continue to next booking instead of failing entire export
            continue
    
    if format_type == 'excel':
        return generate_excel(booking_data)
    elif format_type == 'pdf':
        return generate_pdf(booking_data)
    elif format_type == 'csv':
        return generate_csv(booking_data)
    else:
        return error_response("Invalid format type", "INVALID_REQUEST")

@bookings_bp.route('/<id>', methods=['GET'])
@token_required
def get_booking(id):
    booking = Booking.objects(id=id, agencyId=g.agency_id).first()
    if not booking:
        return not_found_error("Booking not found")
    
    # Manually populate expanded references for UI
    data = mongo_to_dict(booking)
    if booking.packageId:
        data['packageName'] = booking.packageId.name
    if booking.customerId:
        data['customerName'] = booking.customerId.fullName
        
    return jsonify(data), 200

@bookings_bp.route('/<id>', methods=['PUT'])
@token_required
def update_booking(id):
    booking = Booking.objects(id=id, agencyId=g.agency_id).first()
    if not booking:
        return not_found_error("Booking not found")
    
    data = request.get_json()
    if not data:
        return error_response("No data provided", "INVALID_REQUEST")
    
    try:
        # Update fields if provided
        if 'packageId' in data:
            package = Package.objects(id=data['packageId'], agencyId=g.agency_id).first()
            if not package:
                return not_found_error("Package not found")
            booking.packageId = package
        
        if 'category' in data:
            booking.category = data['category']
        
        if 'totalAmount' in data:
            booking.baseAmount = data['totalAmount']
        
        if 'discount' in data:
            booking.discount = data.get('discount', 0)
        
        if 'finalAmount' in data:
            booking.totalAmount = data['finalAmount']
        
        if 'paidAmount' in data:
            booking.paidAmount = data.get('paidAmount', 0)
        
        booking.save()
        
        return jsonify(mongo_to_dict(booking)), 200
    
    except Exception as e:
        return error_response(str(e), "SERVER_ERROR", 500)

def generate_csv(data):
    try:
        if not data:
            return error_response("No data to export", "SERVER_ERROR", 400)
            
        output = io.StringIO()
        headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
        
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=bookings_{datetime.now().strftime('%Y%m%d')}.csv"
        response.headers["Content-type"] = "text/csv"
        return response
    except Exception as e:
        return error_response(f"CSV generation failed: {str(e)}", "SERVER_ERROR", 500)

def generate_excel(data):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Headers
        headers = list(data[0].keys()) if data else []
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_data in data:
            ws.append(list(row_data.values()))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'bookings_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    except ImportError:
        return error_response("Excel export library not installed", "SERVER_ERROR", 500)
    except Exception as e:
        return error_response(f"Excel generation failed: {str(e)}", "SERVER_ERROR", 500)

def generate_pdf(data):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter))
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("<b>Bookings Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = []
        if data:
            # Headers
            headers = list(data[0].keys())
            table_data.append(headers)
            
            # Rows
            for row in data:
                table_data.append(list(row.values()))
        
        # Create table
        if table_data:
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)
        
        doc.build(elements)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'bookings_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    except ImportError:
        return error_response("PDF export library not installed", "SERVER_ERROR", 500)
    except Exception as e:
        return error_response(f"PDF generation failed: {str(e)}", "SERVER_ERROR", 500)
